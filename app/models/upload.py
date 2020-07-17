from app import app, os, celery, ServiceUnavailable, SecurityError, logging

from .resumable import Resumable
from werkzeug.utils import secure_filename

from .record import Record, Property, Trait, Condition, Curve, TableRowParser
import copy
#import grp
from app.cypher import Cypher
from app.emails import send_email
from app.models.parsers import Parsers
from flask import render_template, url_for
from werkzeug import urls
from .user import User
from .queries import Query
from .neo4j_driver import get_driver, bolt_result
import csv
import datetime
import itertools
import contextlib

from openpyxl import load_workbook
from zipfile import BadZipfile


class Upload:
	def __init__(self, username, submission_type, raw_filename):
		self.username = username
		self.submission_type = submission_type
		self.raw_filename = raw_filename
		self.filename = secure_filename(raw_filename)
		self.file_extension = self.filename.rsplit('.', 1)[1].lower()
		self.file_path = os.path.join(app.config['UPLOAD_FOLDER'], username, self.filename)
		self.max_errors = 50
		if self.file_extension == 'csv':
			if self.submission_type == 'table':
				self.file_parser = CSVTable(username, submission_type, raw_filename)

		elif self.file_extension == 'xlsx':
			self.file_parser = UploadXLSX(username, submission_type, raw_filename)


		self.record_types = []
		self.trimmed_file_paths = {}
		self.parse_results = {}
		self.submission_results = {}
		self.fieldnames = {}
		self.contains_data = None
		self.input_variables = {}
		self.row_count = {}
		self.error_messages = {}
		self.property_updater = None


	def check_format(self):
		self.file_parser.check_format()





	def check_duplicate_fieldnames(self, worksheet):
		fieldnames_set = set([i for i in self.fieldnames[worksheet]])
		if len(self.fieldnames[worksheet]) > len(fieldnames_set):
			if self.file_extension == 'xlsx':
				if worksheet in app.config['WORKSHEET_NAMES']:
					error_message = '<p> %s worksheet ' %  app.config['WORKSHEET_NAMES'][worksheet]
				else:
					error_message = '<p> %s worksheet ' % worksheet
			else:
				error_message = '<p>This file '
			self.error_messages.append(error_message + 'contains duplicate column labels. This is not supported.</p>')

	def check_required_fieldnames(self, worksheet):
		record_type = self.get_record_type_from_worksheet(worksheet)
		if record_type:
			if not self.required_sets[record_type].issubset(self.fieldnames[worksheet]):
				self.error_messages.append(
					worksheet +
					' does not appear to contain a full set of required fieldnames for any record type: ' +
					' <br> Property records require: ' + ', '.join(self.required_sets['property']) +
					' <br> Trait and Curve records require: ' + ', '.join(self.required_sets['trait']) +
					' <br> Condition records require:' + ', '.join(self.required_sets['condition'])
				)

	def check_input_variables(self, worksheet):
		# get fieldnames that aren't in the required list or optional list
		# these should be the input variables which we then confirm are found in the db.
		# except for curves, where they should all be numbers
		record_type = self.get_record_type_from_worksheet(worksheet)
		reference_fieldnames = self.required_sets[record_type].union(self.optional_sets[record_type])
		self.input_variables[worksheet] = [
			i for i in self.fieldnames[worksheet] if i.lower() not in reference_fieldnames
		]
		statement = (
			' UNWIND $input_variables AS input_variable '
			'	OPTIONAL MATCH '
			'		(input: Input {name_lower: toLower(trim(toString(input_variable)))}) '
			'	OPTIONAL MATCH '
			'		(input)-[:OF_TYPE]->(record_type: RecordType) '
			'	WITH '
			'		input_variable, record_type '
			'	WHERE '
			'		input IS NULL '
			'		OR '
			'		record_type.name_lower <> $record_type '
			'	RETURN '
			'		input_variable, record_type.name_lower '
		)
		parameters = {
			'record_type': record_type,
			'input_variables': self.input_variables[worksheet]
		}
		with get_driver().session() as neo4j_session:
			unrecognised_input_variables = neo4j_session.read_transaction(
				bolt_result,
				statement,
				parameters
			)
			if unrecognised_input_variables.peek():
				if self.file_extension == 'xlsx':
					error_message = '<p>' + worksheet + ' worksheet '
				else:
					error_message = '<p>This file '
				error_message += (
					'contains column headers that are not recognised as input variables or required details: </p>'
				)
				for input_variable in unrecognised_input_variables :
					error_message += '<dt>' + input_variable[0] + ':</dt> '
					input_record_type = input_variable[1]
					if input_record_type:
						if record_type == 'mixed':
							error_message += (
								' <dd> Required fields missing: '
							)
						else:
							error_message += (
									' <dd> Required fields present for ' +
									record_type +
									' records but this input variable is a ' +
									input_record_type +
									'.'
							)
						if input_record_type == 'condition':
							error_message += (
								'. Condition records require "start date", "start time", "end date" and "end time" '
								' in addition to the "UID" and "Person" fields.'
							)
						elif input_record_type == 'trait':
							error_message += (
								'. Trait records require "date" and "time" fields'
								' in addition to the "UID" and "Person" fields.'
							)
						elif input_record_type == 'property':
							error_message += '. Property records require the "UID" and "Person" fields.'
						error_message += (
							'</dd>\n'
						)
					else:
						error_message += '<dd>Unrecognised input variable. Please check your spelling.</dd>\n'
				self.error_messages.append(error_message)

	def check_fieldnames(self):
		if self.submission_type == 'internal_db':
			pass
		elif not self.fieldnames:
			self.error_messages.append(
				'This file does not contain recognised input'
			)
		else:
			for worksheet in list(self.fieldnames.keys()):
				self.check_duplicate_fieldnames(worksheet)
				self.check_required_fieldnames(worksheet)

	def set_record_types_and_fieldnames_from_csv(self):
		record_type = None
		with open(self.file_path) as uploaded_file:
			file_dict = DictReaderInsensitive(uploaded_file)
			if self.submission_type == 'internal_db':
				record_type = 'mixed'
			elif self.submission_type == 'db':  # this type is submitted through the Correct page
				record_type = 'mixed'
				if not app.config['RECORD_TYPE_REQUIRED'][record_type].issubset(set(file_dict.fieldnames)):
					self.error_messages.append(
						'This file does not appear to be a database exported csv file. ' +
						'It should contain the following fieldnames: ' +
						', '.join([str(i) for i in app.config['RECORD_TYPE_REQUIRED'][record_type]])
					)
			elif self.submission_type == 'table':
				# Consider key fieldnames to set the expected type of input and gather list of input variables
				other_fields = (
						set(file_dict.fieldnames) -
						app.config['RECORD_TYPE_REQUIRED'][record_type] -
						app.config['RECORD_TYPE_OPTIONAL'][record_type]
				)
				if 'start date' in file_dict.fieldnames:
					record_type = 'condition'
				elif 'date' in file_dict.fieldnames:
					# check if all other fieldnames are valid numbers,
					# if so it should be a table for a curve input variable
					all_numbers = True
					for field in other_fields:
						try:
							float(field)
						except ValueError:
							all_numbers = False
							break
					if all_numbers:
						record_type = 'curve'
					else:
						record_type = 'trait'

				else:
					record_type = 'property'
			else:
				self.error_messages.append('Submission type not recognised')
				logging.error(
					'A CSV file was submitted but the type is not recognised to set record type and fieldnames'
				)
		if record_type:
			self.fieldnames = {record_type: file_dict.fieldnames}
			self.record_types = [record_type]
		return record_type

	def load_xlsx(self):
		try:
			wb = load_workbook(self.file_path, read_only=True, data_only=True)
			return wb
		except BadZipfile:
			logging.info(
				'Bad zip file submitted: \n'
				+ 'username: ' + self.username
				+ 'filename: ' + self.file_path
			)
			self.error_messages.append(
				'This file does not appear to be a valid xlsx file'
			)
			return None

	def worksheet_to_curve_input(self, worksheet):
		statement = (
			' MATCH '
			'	(input: Input {name_lower: $name})-[:OF_TYPE]->(:RecordType {name_lower:"curve"}) '
			' RETURN input.name '
		)
		parameters = {
			'name': worksheet.lower()
		}
		with get_driver().session() as neo4j_session:
			result = neo4j_session.read_transaction(
				bolt_result,
				statement,
				parameters
			)
		if result.peek():
			return result.single()[0]
		else:
			self.error_messages.append(
				' This workbook contains an unsupported worksheet:<br> "' + worksheet + '"<br><br>'
				+ 'Other than worksheets named after curve input variables, only the following are accepted: '
				+ '<ul><li>'
				+ '</li><li>'.join(
					[
						str(i) for i in list(app.config['WORKSHEET_NAMES'].values()) if
						i not in app.config['REFERENCE_WORKSHEETS']
					]
				)
				+ '</li>'
			)
			return None

	def check_curve_table(self, worksheet):
		record_type = 'curve'
		reference_fieldnames = self.required_sets[record_type].union(self.optional_sets[record_type])
		other_fields = set(self.fieldnames[worksheet]) - reference_fieldnames
		all_numbers = True
		for field in other_fields:
			try:
				float(field)
			except ValueError:
				all_numbers = False
				break
		if not all_numbers:
			if self.file_extension == 'xlsx':
				error_message = '<p>' + worksheet + ' worksheet'
			else:
				error_message = self.raw_filename
			error_message += (
				' contains unexpected column labels. '
				' For curve data only float values are accepted in addition to the following labels: '
			)
			# we are calling trait required fieldnames here because they are the same for traits and curves
			# and when uploading a csv we check if this is curve data by all other headers not in required list being numbers
			# so the required fieldnames will have been set as trait
			for i in reference_fieldnames:
				error_message += '<dd>' + str(i) + '</dd>'
			self.error_messages.append(error_message)

	def get_record_type(self, worksheet=None):
		record_type = None
		if self.submission_type == 'db':
			record_type = 'mixed'

		if self.
			# page is filename for csv and worksheet for xlsx files

	def get_record_type_from_worksheet(self, worksheet):
		record_type = None
		if self.submission_type == 'db':
			record_type = 'mixed'
		elif worksheet.lower() in app.config['WORKSHEET_TYPES']:
			worksheet_type = app.config['WORKSHEET_TYPES'][worksheet.lower()]
			record_type = None if worksheet_type in app.config['REFERENCE_WORKSHEETS'] else worksheet_type
		else:
			input_variable = self.worksheet_to_curve_input(worksheet)
			if input_variable:
				record_type = 'curve'
				self.input_variables[worksheet] = input_variable
		if record_type and record_type not in self.record_types: # consider using a set for record types, is order important?
			self.record_types.append(record_type)
		return record_type

	def set_record_types_and_fieldnames_from_xlsx(self):
		wb = self.load_xlsx()
		if wb:
			for worksheet in wb.sheetnames:
				record_type = self.get_record_type_from_worksheet(worksheet)
				if record_type:
					ws = wb[worksheet]
					rows = ws.iter_rows(min_row=1, max_row=1)
					first_row = next(rows)
					self.fieldnames[worksheet] = [
						c.value.strip().lower()
						if isinstance(c.value, str)
						else str(c.value)
						for c in first_row
						if c.value
					]
					if record_type == 'curve':
						# check the fieldnames that aren't required/optional for curved inputs are all numbers:
						self.check_curve_table(worksheet)
					else:
						self.check_input_variables(worksheet)

	def set_record_types_and_fieldnames(self):
		if self.file_extension == 'csv':
			self.set_record_types_and_fieldnames_from_csv()
		elif self.file_extension == 'xlsx':
			self.set_record_types_and_fieldnames_from_xlsx()
		else:
			self.error_messages.append('Only csv and xlsx file formats are supported for data submissions')
		self.check_fieldnames()

	def process_file(self):
		self.set_record_types_and_fieldnames()
		if self.error_messages:
			return
		if self.file_extension == 'csv':
			self.process_csv()
		elif self.file_extension == 'xlsx':
			self.process_xlsx()

	def process_csv(self):
		max_rows = app.config['BATCH_PROCESS_ROW_COUNT']
		with open(self.file_path) as file:
			file_dict = csv.DictReader(file)
			record_type = self.record_types[0]
			page_name = record_type


			record_batch = RecordBatch(self.username, page_name, record_type)
			for i, row in enumerate(file_dict, start=1):
				if i == max_rows:
					record_batch.row_end = i
					record_batch.submit()
				record_batch.add_row(row, i, record_type)




				if len(self.error_messages) >= self.max_errors:
					break

					row['row_index'] = i









			self.row_count[self.filename] = i





	def process_worksheet(self, file_path, worksheet):
		wb = load_workbook(self.file_path, read_only=True, data_only=True)




	# clean up the csv by passing through dict reader and rewriting
	def trim_file(self, worksheet):
		trimmed_file_path = '_'.join([
			os.path.splitext(self.file_path)[0],
			worksheet,
			'trimmed.csv'
		])
		self.trimmed_file_paths[worksheet] = trimmed_file_path
		if self.file_extension == 'csv':
			with open(self.file_path) as uploaded_file, \
					open(os.open(trimmed_file_path, os.O_CREAT | os.O_WRONLY | os.O_TRUNC, 0o640), "w") as trimmed_file:
				# this dict reader lowers case and trims whitespace on all headers
				file_dict = DictReaderInsensitive(
					uploaded_file,
					skipinitialspace=True
				)
				file_writer = csv.DictWriter(
					trimmed_file,
					fieldnames=['row_index'] + file_dict.fieldnames,
					quoting=csv.QUOTE_ALL
				)
				file_writer.writeheader()
				for i, row in enumerate(file_dict):
					self.row_count[worksheet] = i + 1
					# remove rows without entries
					if any(field.strip() for field in row):
						for field in file_dict.fieldnames:
							if row[field]:
								row[field] = row[field].strip()
						row['row_index'] = self.row_count[worksheet]
						file_writer.writerow(row)
		elif all([
			self.file_extension == 'xlsx',
			self.submission_type in ['db', 'table']
		]):
			wb = load_workbook(self.file_path, read_only=True, data_only=True)
			if self.submission_type == 'db':
				ws = wb['Records']
			elif self.submission_type == 'table':
				if worksheet in app.config['WORKSHEET_NAMES']:
					ws = wb[app.config['WORKSHEET_NAMES'][worksheet]]
				else:
					ws = wb[worksheet]
			# here opening trimmed file with permissions allowing the group (neo4j) to have read access
			with open(os.open(trimmed_file_path, os.O_CREAT | os.O_WRONLY | os.O_TRUNC, 0o640), "w") as trimmed_file:
				file_writer = csv.writer(
					trimmed_file,
					quoting=csv.QUOTE_ALL
				)
				rows = ws.iter_rows()
				first_row = next(rows)
				file_writer.writerow(
					['row_index'] +
					[cell.value.lower() if isinstance(cell.value, str) else str(cell.value) for cell in first_row if cell.value]
				)
				# handle deleted columns in the middle of the worksheet
				empty_headers = []
				time_column_index = None
				date_column_index = None
				for i, cell in enumerate(first_row):
					if not cell.value:
						empty_headers.append(i)
					if isinstance(cell.value, str):
						if cell.value.strip().lower() == 'time':
							time_column_index = i
				for j, row in enumerate(rows):
					# j+2 to store the "Line number" as index
					# this is 0 based, and accounts for header
					self.row_count[worksheet] = j + 2
					cell_values = [cell.value for cell in row]
					# remove columns with empty header
					for i in sorted(empty_headers, reverse=True):
						del cell_values[i]
					# remove empty rows
					for i, value in enumerate(cell_values):
						if isinstance(value, datetime.datetime):
							# when importing from excel if the time isn't between 00:00 and 24:00
							# then it is imported as a datetime object relative to 1900,1,1,0,0
							# so we catch this here by checking if we are looking at the time column
							if i == time_column_index:
								cell_values[i] = value.strftime("%H:%M")
							else:
								try:
									cell_values[i] = value.strftime("%Y-%m-%d")
								except ValueError:
									cell_values[i] = 'Dates before 1900 are not supported'
						elif isinstance(value, datetime.time):
							cell_values[i] = value.strftime("%H:%M")
						else:
							if isinstance(value, str):
								cell_values[i] = value.strip()
					if any(value for value in cell_values):
						cell_values = [self.row_count[worksheet]] + cell_values
						file_writer.writerow(cell_values)

	def parse_rows(self, worksheet):
		trimmed_file_path = self.trimmed_file_paths[worksheet]
		submission_type = self.submission_type
		parse_result = ParseResult(submission_type, worksheet, self.fieldnames[worksheet])
		self.parse_results[worksheet] = parse_result
		if submission_type == 'table':
			record_type = self.get_record_type_from_worksheet(worksheet)
		elif submission_type == 'db':
			record_type = 'mixed'
		if record_type == 'curve':
			x_values = set(self.fieldnames[worksheet]) - self.required_sets[record_type]
		with open(trimmed_file_path) as trimmed_file:
			trimmed_dict = csv.DictReader(trimmed_file)
			for row in trimmed_dict:
				if submission_type == 'table':
					if record_type != 'curve':
						# first check for input data, if none then just skip this row
						if worksheet in self.input_variables:
							if [row[input_variable] for input_variable in self.input_variables[worksheet] if row[input_variable]]:
								parse_result.contains_data = True
								parse_result.parse_table_row(row)
					else:
						if [row[x] for x in x_values if row[x]]:
							parse_result.contains_data = True
							parse_result.parse_table_row(row)
				elif submission_type == 'db':
					parse_result.parse_db_row(row)
		if parse_result.contains_data:
			self.contains_data = True
		if parse_result.duplicate_keys:
			self.error_messages.append(
				'<p>' + worksheet
				+ 'worksheet contains duplicate keys:</p>'
				+ parse_result.duplicate_keys_table()
			)

	def db_check(
		self,
		neo4j_session,
		worksheet
	):
		# todo it seems like the field errors here should not occur
		# todo we are handling this check before we parse the rows
		# todo so we could remove this check for input variables from  here to simplify
		username = self.username
		submission_type = self.submission_type
		if submission_type == 'db':
			record_type = 'mixed'
		elif worksheet.lower() in app.config['WORKSHEET_TYPES']:
			record_type = app.config['WORKSHEET_TYPES'][worksheet.lower()]
		else:
			record_type = 'curve'
		trimmed_file_path = self.trimmed_file_paths[worksheet]
		trimmed_filename = os.path.basename(trimmed_file_path)
		parse_result = self.parse_results[worksheet]
		with open(trimmed_file_path) as trimmed_file:
			if submission_type == 'db':
				trimmed_dict_reader = DictReaderInsensitive(trimmed_file)
				inputs_set = set()
				# todo move this to the parse procedure where we iterate through the file already
				for row in trimmed_dict_reader:
					inputs_set.add(row['input variable'].lower())
				record_types = neo4j_session.run(
					' UNWIND $inputs as input_name'
					'	MATCH '
					'	(f:Input { '
					'		name_lower: input_name '
					'	})-[:OF_TYPE]->(record_type: RecordType) '
					' RETURN distinct(record_type.name_lower) ',
					inputs=list(inputs_set)
				).value()
				if not set(self.required_sets[record_type]).issubset(set(self.fieldnames[worksheet])):
					missing_fieldnames = set(self.required_sets[record_type]) - set(self.fieldnames[worksheet])
					if self.file_extension == 'xlsx':
						error_message = '<p>' + app.config['WORKSHEET_NAMES'][worksheet] + ' worksheet '
					else:
						error_message = '<p>This file '
					error_message += (
							' is missing the following required fields: '
							+ ', '.join([i for i in missing_fieldnames])
							+ '</p>'
					)
					self.error_messages.append(error_message)
			elif submission_type == 'table' and worksheet in self.input_variables:
				if record_type == 'property':
					statement = Cypher.upload_table_property_check
					parameters = {
						'username': username,
						'filename': urls.url_fix('file:///' + username + '/' + trimmed_filename ),
						'inputs': self.input_variables[worksheet],
						'record_type': record_type
					}
				elif record_type == 'trait':
					statement = Cypher.upload_table_trait_check
					parameters = {
						'username': username,
						'filename': urls.url_fix('file:///' + username + '/' + trimmed_filename),
						'inputs': self.input_variables[worksheet],
						'record_type': record_type
					}
				elif record_type == 'condition':
					statement = Cypher.upload_table_condition_check
					parameters = {
						'username': username,
						'filename': urls.url_fix('file:///' + username + '/' + trimmed_filename),
						'inputs': self.input_variables[worksheet],
						'record_type': record_type
					}
				elif record_type == 'curve':
					statement = Cypher.upload_table_curve_check
					reference_fields = self.required_sets[record_type].union(self.optional_sets[record_type])
					parameters = {
						'username': username,
						'filename': urls.url_fix('file:///' + username + '/' + trimmed_filename),
						'input_name': self.input_variables[worksheet],
						'x_values': sorted(
							[float(i) for i in self.fieldnames[worksheet] if i.lower() not in reference_fields]
						),
						'record_type': record_type
					}
				else:
					logging.warn('record type not recognised')
					return
				check_result = neo4j_session.run(
					statement,
					parameters
				)
				trimmed_dict_reader = DictReaderInsensitive(trimmed_file)
				row = next(trimmed_dict_reader)
				# need to check_result sorted by file/dictreader row_index
				for item in check_result:
					record = item[0]
					while record['row_index'] != int(row['row_index']):
						row = next(trimmed_dict_reader)
					if not record['UID']:
						parse_result.merge_error(
							row,
							"uid",
							"missing"
						)
					if not record['Input variable']:
						parse_result.add_field_error(
							record['Supplied input name'],
							(
								"This input variable is not found. Please check your spelling. "
								"This may also be because the input variable is not available at the level of these items"
							)
						)
					# we add found fields to a list to handle mixed items in input
					# i.e. if found at level of one item but not another
					else:
						parse_result.add_field_found(record['Input variable'])
					if all([
						record['UID'],
						record['Input variable'],
						not record['Value']
					]):
						parse_result.merge_error(
							row,
							record['Supplied input name'],
							"format",
							input_name=record['Input variable'],
							input_format=record['Format'],
							category_list=record['Category list']
						)
					# need to check an element of the list as all results
					if record['Conflicts'][0]['existing_value']:
						parse_result.merge_error(
							row,
							record['Supplied input name'],
							"conflict",
							conflicts=record['Conflicts']
						)
				if parse_result.field_found:
					for field in parse_result.field_found:
						parse_result.rem_field_error(field)
		if parse_result.field_errors:
			if self.file_extension != 'xlsx':
				self.error_messages.append(
					'<p> This file contains unrecognised column labels: </p><ul><li>'
					+ '</li><li>'.join(parse_result.field_errors)
					+ '</li></ul>'
				)
			else:
				if worksheet in app.config['WORKSHEET_NAMES']:
					self.error_messages.append(
						'<p>' + app.config['WORKSHEET_NAMES'][worksheet]
						+ ' worksheet contains unrecognised column labels: </p><ul><li>'
						+ '</li><li>'.join(parse_result.field_errors)
						+ '</li></ul>'
					)
				else:
					self.error_messages.append(
						'<p>' + worksheet
						+ ' worksheet contains unrecognised column labels: </p><ul><li>'
						+ '</li><li>'.join(parse_result.field_errors)
						+ '</li></ul>'
					)
		if parse_result.errors:
			if self.file_extension != 'xlsx':
				self.error_messages.append(
					'<p> This file contains errors: </p>'
					+ parse_result.html_table()
				)
			else:
				if worksheet in app.config['WORKSHEET_NAMES']:
					self.error_messages.append(
						'<p>' + app.config['WORKSHEET_NAMES'][worksheet]
						+ ' worksheet contains errors: '
						+ parse_result.html_table()
					)
				else:
					self.error_messages.append(
						'<p>' + worksheet
						+ ' worksheet contains errors: '
						+ parse_result.html_table()
					)

	def submit(
			self,
			neo4j_session,
			worksheet
	):
		username = self.username
		if worksheet.lower() in app.config['WORKSHEET_TYPES']:
			record_type = app.config['WORKSHEET_TYPES'][worksheet.lower()]
		else:
			record_type = 'curve'
		trimmed_file_path = self.trimmed_file_paths[worksheet]
		trimmed_filename = os.path.basename(trimmed_file_path)
		submission_type = self.submission_type
		filename = self.filename
		inputs = self.input_variables[worksheet] if worksheet in self.input_variables else None
		with contextlib.closing(SubmissionResult(username, filename, submission_type, record_type)) as submission_result:
			self.submission_results[worksheet] = submission_result
			if record_type == 'property':
				statement = Cypher.upload_table_property
				result = neo4j_session.run(
					statement,
					username=username,
					filename=urls.url_fix("file:///" + username + '/' + trimmed_filename),
					inputs=inputs,
					record_type=record_type
				)
			elif record_type == 'trait':
				statement = Cypher.upload_table_trait
				result = neo4j_session.run(
					statement,
					username=username,
					filename=urls.url_fix("file:///" + username + '/' + trimmed_filename),
					inputs=inputs,
					record_type=record_type
				)
			elif record_type == 'condition':
				statement = Cypher.upload_table_condition
				result = neo4j_session.run(
					statement,
					username=username,
					filename=urls.url_fix("file:///" + username + '/' + trimmed_filename),
					inputs=inputs,
					record_type=record_type
				)
			elif record_type == 'curve':
				statement = Cypher.upload_table_curve
				reference_fields = self.required_sets[record_type].union(self.optional_sets[record_type])
				result = neo4j_session.run(
					statement,
					username=username,
					filename=urls.url_fix("file:///" + username + '/' + trimmed_filename),
					input_name=self.input_variables[worksheet],
					x_values=sorted(
						[float(i) for i in self.fieldnames[worksheet] if i not in reference_fields]
					),
					record_type=record_type
				)
			else:
				logging.warning('Record type not recognised')
			# create a submission result and update properties from result
			if record_type == 'property':
				self.property_updater = PropertyUpdateHandler(neo4j_session)
			for record in result:
				logging.debug('Now parse result and update properties in the graph')
				submission_result.parse_record(record[0])
				if record_type == 'property':
					if self.property_updater.process_record(record[0]):
						break
			# As we are collecting property updates we need to run the updater at the end
			if not result.peek():
				if record_type == 'property':
					self.property_updater.update_all()
					self.property_updater.tx.commit()

	def error_response(self):
		with app.app_context():
			html = render_template(
				'emails/upload_report.html',
				response='<br>'.join(self.error_messages)
			)
			subject = "BreedCAFS upload rejected"
			recipients = [User(self.username).find('')['email']]
			response = "Submission rejected due to invalid file:\n " + '<br>'.join(self.error_messages)
			body = response
			send_email(subject, app.config['ADMINS'][0], recipients, body, html)
		return {
			'status': 'ERRORS',
			'result': (
					'Submission report for file: ' + self.raw_filename + '<br><br>' +
					'<br>'.join(self.error_messages)
			)
		}

	@celery.task(bind=True)
	def async_submit(self, username, upload_object):
		try:
			if upload_object.file_extension not in ['csv', 'xlsx']:
				return {
					'status': 'SUCCESS',
					'result': "Nothing done with the file"
				}

			upload_object.check_file_format()
			if upload_object.error_messages:
				return upload_object.error_response()
			upload_object._file()







			if upload_object.error_messages:
				with app.app_context():
					html = render_template(
						'emails/upload_report.html',
						response='<br>'.join(upload_object.error_messages)
					)
					subject = "BreedCAFS upload rejected"
					recipients = [User(username).find('')['email']]
					response = "Submission rejected due to invalid file:\n " + '<br>'.join(upload_object.error_messages)
					body = response
					send_email(subject, app.config['ADMINS'][0], recipients, body, html)
				return {
					'status': 'ERRORS',
					'result': (
							'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
							'<br>'.join(upload_object.error_messages)
					)
				}
			logging.debug('Start database session')
			with get_driver().session() as neo4j_session:


				for worksheet in list(upload_object.fieldnames.keys()):
					upload_object.process_page

					# clean up the file removing empty lines and whitespace, lower case headers for matching in db
					upload_object.trim_file(worksheet)
					# parse the trimmed file/worksheet for errors
					# also adds parse_result to upload_object.parse_result dict (by record_type)
					upload_object.parse_rows(worksheet)
					# with string parsing performed, now we check against the database for UID, input variable, value
					upload_object.db_check(neo4j_session, worksheet)
				if not upload_object.contains_data:
					upload_object.error_messages.append(
						upload_object.raw_filename
						+ ' appears to contain no input values. '
					)
				if upload_object.error_messages:
					error_messages = '<br>'.join(upload_object.error_messages)
					with app.app_context():
						html = render_template(
							'emails/upload_report.html',
							response=(
								'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
								error_messages
							)
						)
						subject = "BreedCAFS upload rejected"
						recipients = [User(username).find('')['email']]
						body = (
								'Submission report for file: ' + upload_object.raw_filename + '<br><br>'
								+ error_messages
						)
						send_email(subject, app.config['ADMINS'][0], recipients, body, html)
					return {
						'status': 'ERRORS',
						'result': (
							'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
							error_messages
						)
					}
				conflict_files = []
				submissions = []
				for worksheet in list(upload_object.fieldnames.keys()):
					# submit data
					upload_object.submit(neo4j_session, worksheet)
					submission_result = upload_object.submission_results[worksheet]
					if submission_result.conflicts_found:
						conflict_files.append(submission_result.conflicts_file_path)
						os.unlink(submission_result.submissions_file_path)
					else:
						os.unlink(submission_result.conflicts_file_path)
						if submission_result.submissions_file_path:
							submissions.append((
								worksheet,
								submission_result.summary(),
								submission_result.submissions_file_path
							))
				if conflict_files:
					# These should only be generated due to concurrent conflicting submissions
					# or internal conflicts in the submitted file
					with neo4j_session.begin_transaction() as tx:
						for worksheet in list(upload_object.fieldnames.keys()):
							deletion_result = upload_object.correct(tx, 'user', worksheet, revert_submission=True)
							property_uid = {
								'set custom name': [],
								'set row': [],
								'set column': [],
								'set sample unit': [],
								'assign tree to block by name': [],
								'assign tree to block by id': [],
								'assign sample to block(s) by name': [],
								'assign sample to block(s) by id': [],
								'assign sample to tree(s) by id': [],
								'assign sample to sample(s) by id': [],
								'set harvest date': [],
								'set planting date': [],
								'set harvest time': [],
								'assign variety name': [],
								'assign variety (el frances code)': []
							}
							for record in deletion_result:
								if record[0]['input variable'] in property_uid:
									property_uid[record[0]['input variable']].append(record[0]['uid'])
					with neo4j_session.begin_transaction() as tx:
						upload_object.remove_properties(tx, property_uid)
					with app.app_context():
						response = (
							'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
							'<p>Conflicting records were either submitted concurrently '
							'or are found within in the submitted file. '
							'Your submission has been rejected. Please address the conflicts listed '
							'in the following files before resubmitting. '
							'\n'
						)
						for conflict_file in conflict_files:
							response += ' - '
							response += (
								url_for(
									'download_file',
									username=username,
									filename=os.path.basename(conflict_file),
									_external=True
								)
							)
							response += '\n'
						response += '</p>'
						html = render_template(
							'emails/upload_report.html',
							response=(
								'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
								response
							)
						)
						# send result of merger in an email
						subject = "BreedCAFS upload rejected"
						recipients = [User(username).find('')['email']]
						body = (
							'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
							response
						)
						send_email(subject, app.config['ADMINS'][0], recipients, body, html)
						return {
							'status': 'ERRORS',
							'result': (
									'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
									response
							)
						}
				if 'property' in upload_object.record_types:
					property_updater = upload_object.property_updater
					if property_updater.errors:
						property_updater.format_error_list()
						error_messages = [
							item for errors in list(property_updater.errors.values()) for item in errors
						]
						error_messages = '<br>'.join(error_messages)
						with neo4j_session.begin_transaction() as tx:
							for worksheet in list(upload_object.fieldnames.keys()):
								deletion_result = upload_object.correct(
									tx,
									'user',
									worksheet,
									revert_submission=True
								)
								property_uid = {
									'set custom name': [],
									'set row': [],
									'set column': [],
									'set sample unit': [],
									'assign tree to block by name': [],
									'assign tree to block by id': [],
									'assign sample to block(s) by name': [],
									'assign sample to block(s) by id': [],
									'assign sample to tree(s) by id': [],
									'assign sample to sample(s) by id': [],
									'set harvest date': [],
									'set planting date': [],
									'set harvest time': [],
									'assign variety name': [],
									'assign variety (el frances code)': []
								}
								for record in deletion_result:
									if record[0]['input variable'] in property_uid:
										property_uid[record[0]['input variable']].append(record[0]['uid'])
						with neo4j_session.begin_transaction() as tx:
							upload_object.remove_properties(tx, property_uid)
						with app.app_context():
							response = (
									'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
									error_messages
							)
							html = render_template(
								'emails/upload_report.html',
								response=response
							)
							subject = "BreedCAFS upload rejected"
							recipients = [User(username).find('')['email']]
							body = response
							send_email(subject, app.config['ADMINS'][0], recipients, body, html)
							return {
								'status': 'ERRORS',
								'result': response
							}
				# now need app context for the following (this is running asynchronously)
				with app.app_context():
					if not submissions:
						response = 'No data submitted, please check that you uploaded a completed file'
						return {
							'status': 'ERRORS',
							'result': (
								'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
								response
							)
						}
					else:
						response = "<p>"
						for submission in submissions:
							submission_url = url_for(
								'download_file',
								username=username,
								filename=os.path.basename(submission[2]),
								_external=True,
								# adding a parameter here to stop browser from accessing cached version of file
								date=datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S-%f')
							)
							if submission[0] in app.config['WORKSHEET_NAMES']:
								response += (
									' <br> - <a href= ' + submission_url + '>'
									+ str(app.config['WORKSHEET_NAMES'][submission[0]])
									+ '</a> contained ' + str(submission[1]['submitted']) + ' new '
									+ ' and ' + str(submission[1]['resubmissions']) + ' existing records.'
								)
							else:
								response += (
										' <br> - <a href= ' + submission_url + '>'
										+ str(submission[0])
										+ '</a> contained ' + str(submission[1]['submitted']) + ' new '
										+ ' and ' + str(submission[1]['resubmissions']) + ' existing records.'
								)
						response += '</p>'
						html = render_template(
							'emails/upload_report.html',
							response=(
								'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
								response
							)
						)
						# send result of merger in an email
						subject = "BreedCAFS upload summary"
						recipients = [User(username).find('')['email']]
						body = (
							'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
							response
						)
						send_email(subject, app.config['ADMINS'][0], recipients, body, html)
						return {
							'status': 'SUCCESS',
							'result': (
								'Submission report for file: ' + upload_object.raw_filename + '<br><br>' +
								response
							)
						}
		except (ServiceUnavailable, SecurityError) as exc:
			raise self.retry(exc=exc)

	def correct(
			self,
			tx,
			access,
			worksheet,
			revert_submission=False
	):
		if access == 'global_admin':
			statement = (
				' MATCH '
				'	(partner:Partner) '
				'	<-[:AFFILIATED {data_shared:True})-(user:User) '
				' MATCH '
				'	(current_user: User { '
				'		username_lower: toLower($username) '
				'	}) '
				' WHERE "global_admin" IN current_user.access '
			)
		elif access == 'partner_admin':
			statement = (
				' MATCH '
				'	(current_user: User {'
				'		username_lower: toLower($username)'
				'	}) '
				'	-[: AFFILIATED {admin: True}]->(partner: Partner) '
				' WHERE "partner_admin" IN current_user.access '
				' MATCH '
				'	(partner) '
				'	<-[:AFFILIATED {data_shared:True}]-(user:User) '
			)
		else:
			statement = (
				' MATCH '
				'	(partner: Partner) '
				'	<-[:AFFILIATED {data_shared:True}]-(user: User {'
				'		username_lower: toLower($username)'
				'	}) '
				' WITH '
				'	partner, '
				'	user as user, '
				'	user as current_user '
			)
		statement += (
			' LOAD CSV WITH HEADERS FROM $filename as csvLine '
			'	WITH '
			'	current_user, '
			'	user, partner, '
			'	toInteger(csvLine.row_index) as row_index, '
			'	apoc.date.parse('
			'		csvLine.`submitted at`,'
			'		"ms", '
			'		 "yyyy-MM-dd HH:mm:ss"'
			'	) as `submitted at`, '
			'	toLower(csvLine.`input variable`) as input_name, '
			'	CASE '
			'		WHEN toInteger(csvLine.uid) IS NOT NULL '
			'		THEN toInteger(csvLine.uid) '
			'		ELSE toUpper(csvLine.uid)'
			'	END as uid, ' 
			'	toInteger(csvLine.replicate) as replicate, '
			'	CASE '
			'		WHEN csvLine.time <> "" '
			'		THEN apoc.date.parse( '
			'			csvLine.time, '
			'			"ms", '
			'			 "yyyy-MM-dd HH:mm" '
			'			) '
			'		ELSE Null '
			'	END as time, '
			'	CASE '
			'		WHEN size(split(csvLine.period, " - ")) > 1 '
			'		THEN CASE '
			'			WHEN toLower(split(csvLine.period, " - ")[0]) = "undefined" '
			'			THEN False '
			'			ELSE apoc.date.parse('
			'				split(csvLine.period, " - ")[0],'
			'				"ms", '
			'				 "yyyy-MM-dd HH:mm" '
			'			) '
			'		END '
			'		ELSE Null '
			'	END as start, '
			'	CASE '
			'		WHEN size(split(csvLine.period, " - ")) > 1 '
			'		THEN CASE '
			'			WHEN toLower(split(csvLine.period, " - ")[1]) = "undefined" '
			'			THEN False '
			'			ELSE apoc.date.parse('
			'				split(csvLine.period, " - ")[1],'
			'				"ms", '
			'				 "yyyy-MM-dd HH:mm"'
			'			) '
			'		END '
			'		ELSE Null '
			'	END as end, '
			'	CASE '
			# match at least one valid entry
			'		WHEN csvLine.value =~ "\\\\[.+,.+\\\\].*" '
			'		THEN '		
			'			[x in split(csvLine.value, "],") | [i in split(x, ",") | toFloat(replace(replace(i, "[", ""), "]", ""))][0]] '
			'		ELSE Null '
			'	END as x_values, '
			'	CASE '
			# match at least one valid entry
			'		WHEN csvLine.value =~ "\\\\[.+,.+\\\\].*" '
			'		THEN '		
			'			[x in split(csvLine.value, "],") | [i in split(x, ",") | toFloat(replace(replace(i, "[", ""), "]", ""))][1]] '
			'		ELSE Null '
			'	END as y_values '		
			' MATCH '
			'	(user)'
			'	-[:SUBMITTED]->(:Submissions)  '
			'	-[:SUBMITTED]->(:Records) '
			'	-[:SUBMITTED]->(uff:UserFieldInput) '
			'	-[submitted:SUBMITTED]->(record :Record) '
			'	-[record_for:RECORD_FOR]->(if:ItemInput) '
			'	-[:FOR_ITEM]-(item:Item {'
			'		uid: uid'
			'	}), '
			'	(if)'
			'	-[FOR_INPUT*..2]->(input:Input {'
			'		name_lower:input_name'
			'	})-[:OF_TYPE]->(record_type:RecordType) '
			' WHERE '
			# account for rounding to nearest second for submitted at in output files
			'	round(submitted.time / 1000) * 1000 = `submitted at` '
			'	AND '
			'	CASE '
			'		WHEN record.time IS NULL '
			'		THEN True '
			'		ELSE record.time = time '
			'	END '
			'	AND '
			'	CASE '
			'		WHEN record.start IS NULL'
			'		THEN True '
			'		ELSE record.start = start '
			'	END'
			'	AND '
			'	CASE '
			'		WHEN record.end IS NULL '
			'		THEN True '
			'		ELSE record.end = end '
			'	END '
			'	AND '
			'	CASE '
			'		WHEN record.replicate IS NULL OR record.replicate = 0 '
			'		THEN True '
			'		ELSE record.replicate = replicate '
			'	END '
			'	AND '
			'	CASE '
			'		WHEN record_type.name_lower = "curve" '
			'		AND record.x_values = x_values '
			'		AND record.y_values = y_values '
			'		THEN True '
			'		ELSE record_type.name_lower IN ["property", "trait", "condition"] '
			'	END '
			' MERGE '
			'	(current_user)'
			'	-[:DELETED]->(del:Deletions) '
			' CREATE '
			'	(del)-[:DELETED { '
			'		`deleted at`: datetime.transaction().epochMillis '
			'	}]->(record) '
			' CREATE '
			'	(record)-[:DELETED_FROM]->(if) '
			' DELETE '
			'	record_for '
			' RETURN { '
			'	time: time, '
			'	record_time: record.time, '
			'	uid: uid, '
			'	`input variable`: input.name_lower, '
			'	row_index: row_index '
			' } '
			' ORDER BY row_index '
		)
		if revert_submission:
			trimmed_filename = self.submission_results[worksheet].new_records_filename
		else:
			trimmed_file_path = self.trimmed_file_paths[worksheet]
			trimmed_filename = os.path.basename(trimmed_file_path)
		result = tx.run(
				statement,
				username=self.username,
				access=access,
				filename=("file:///" + self.username + '/' + trimmed_filename)
			)
		#if revert_submission:
			#os.unlink(self.submission_results[worksheet].new_records_file_path)
		return result

	@staticmethod
	def remove_properties(tx, property_uid):
		inputs_properties = app.config['INPUTS_PROPERTIES']
		properties_inputs = app.config['PROPERTIES_INPUTS']
		for input_variable, item_uids in property_uid.items():
			if property_uid[input_variable]:
				if inputs_properties[input_variable] == 'name':
					tx.run(
						' UNWIND $uid_list as uid'
						' MATCH '
						'	(item: Item {uid: uid}) '
						' REMOVE item.name ',
						uid_list=property_uid[input_variable]
					)
				elif inputs_properties[input_variable] == 'row':
					tx.run(
						' UNWIND $uid_list as uid'
						' MATCH '
						'	(item: Item {uid: uid}) '
						' REMOVE item.row ',
						uid_list=property_uid[input_variable]
					)
				elif inputs_properties[input_variable] == 'column':
					tx.run(
						' UNWIND $uid_list as uid'
						' MATCH '
						'	(item: Item {uid: uid}) '
						' REMOVE item.column ',
						uid_list=property_uid[input_variable]
					)
				elif inputs_properties[input_variable] == 'location':
					tx.run(
						' UNWIND $uid_list as uid'
						' MATCH '
						'	(item: Item {uid: uid}) '
						' REMOVE item.location ',
						uid_list=property_uid[input_variable]
					)
				elif inputs_properties[input_variable] == 'variety':
					tx.run(
						' UNWIND $uid_list as uid '
						' MATCH '
						'	(item: Item { '
						'		uid: uid '
						'	}) '
						' OPTIONAL MATCH '
						'	(item) '
						'	<-[:FOR_ITEM]-(ii: ItemInput) '
						'	-[:FOR_INPUT*2]->(i: Input), '
						'	(ii)<-[:RECORD_FOR]-(:Record) '
						' WHERE i.name_lower IN $property_inputs '
						' WITH '
						'	item '
						' WHERE ii IS NULL '
						' MATCH '
						'	(item)-[of_variety:OF_VARIETY]->(fv:FieldVariety) '
						'	<-[contains_variety:CONTAINS_VARIETY]-(field:Field) '
						' DELETE of_variety '
						' REMOVE item.variety '
						' WITH '
						'	item, fv, contains_variety '
						'	OPTIONAL MATCH '
						'		(item)-[:IS_IN | FROM *]->(ancestor: Item) '
						' WITH '
						'	item, fv, contains_variety, '
						'	collect(distinct ancestor) as ancestors '
						' OPTIONAL MATCH '
						'	(item)<-[:IS_IN | FROM *]-(descendant: Item) '
						' WITH  '
						'	item, fv, contains_variety, '
						'	ancestors + collect(distinct descendant) as lineage '
						' UNWIND lineage AS kin '
						'	OPTIONAL MATCH '
						'		(kin)-[:IS_IN | FROM *]->(kin_ancestor: Item) '
						'	WITH '
						'		item, fv, contains_variety, '
						'		kin, '
						'		collect(distinct kin_ancestor) as kin_ancestors '
						'	OPTIONAL MATCH '
						'		(kin)<-[:IS_IN | FROM *]-(kin_descendant: Item) '
						'	WITH '
						'		item, fv, contains_variety, '
						'		kin, '			
						'		kin_ancestors + collect(distinct kin_descendant) as kin_lineage '
						'	UNWIND '
						'		kin_lineage as kin_of_kin '
						'	WITH '
						'		item, fv, contains_variety, '
						'		kin, '
						'		collect(distinct kin_of_kin.variety) as kin_varieties '
						'	SET kin.varieties = CASE WHEN kin.variety IS NOT NULL THEN [kin.variety] ELSE kin_varieties END '
						' WITH '
						'	item, fv, contains_variety, '
						'	kin, '
						'	collect(kin.variety) as varieties '
						' SET item.varieties = CASE WHEN item.variety IS NOT NULL THEN [item.variety] ELSE varieties END '
						' WITH DISTINCT '
						'	fv, contains_variety '
						' OPTIONAL MATCH '
						'	(:Item)-[of_variety:OF_VARIETY]->(fv) '
						' FOREACH (n IN CASE WHEN of_variety IS NULL THEN [1] ELSE [] END | '
						'	DELETE contains_variety '
						' ) ',
						uid_list=property_uid[input_variable],
						property_inputs=properties_inputs[inputs_properties[input_variable]]
					)
				elif inputs_properties[input_variable] == 'unit':
					# There are multiple inputs for unit
					# we only need to update the property if there is no other record
					tx.run(
						' UNWIND $uid_list as uid'
						' MATCH '
						'	(item: Item {uid: uid}) '
						' OPTIONAL MATCH '
						'	(item) '
						'	<-[:FOR_ITEM]-(ii: ItemInput) '
						'	-[:FOR_INPUT*2]->(i: Input), '
						'	(ii)<-[:RECORD_FOR]-(:Record) '
						' WHERE i.name_lower IN $property_inputs '
						' WITH item WHERE ii IS NULL '
						' REMOVE item.unit ',
						uid_list=property_uid[input_variable],
						property_inputs=properties_inputs[inputs_properties[input_variable]]
					)
				elif inputs_properties[input_variable] == 'source':
					# There may be multiple inputs for tree/sample item sources (by id, by name)
					#  and multiple contributions with variable specificity for samples
					#  but conflicts are prevented between assignments at same level and
					#  only assignment to greater specificity is allowed so
					#  we only need to update if the deleted record is the latest "source" submission for this item
					#  we could check this by comparing the latest record 'value' against the current assignments
					#  but it might be sensible to just always update to assignment from latest record submission
					# we do also need to update the varieties for new item lineage and prior ancestors
					statement = (
						' UNWIND $uid_list as uid '
						' MATCH '
						'	(item: Item {uid: uid}) '
						'	-[:FROM | IS_IN*]->(ancestor:Item) '
						' WITH item, collect(ancestor) as prior_ancestors '
						' OPTIONAL MATCH '
						'	(item) '
						'	<-[:FOR_ITEM]-(ii: ItemInput) '
						'	-[:FOR_INPUT*2]->(input: Input), '
						'	(ii)<-[:RECORD_FOR]-(record:Record) '
						'	<-[s:SUBMITTED]-() '
						' WHERE input.name_lower IN $property_inputs '
					)
					if 'to block' in input_variable:
						# these will revert to field level if no records
						# but remain unchanged if existing source assignment records
						statement += (
							' AND record IS NULL '
						)
						if 'assign tree' in input_variable:
							# these are trees so block relationship can just be deleted and the counter decremented
							# The IS_IN field relationship is retained when assigning tree to block
							statement += (
								' WITH item, prior_ancestors '
								' MATCH (item)-[is_in:IS_IN]->(:BlockTrees)<-[:FOR]-(c:Counter) '
								' SET c._LOCK_ = True '
								' DELETE is_in '
								' SET c.count = c.count - 1 '
								' REMOVE c._LOCK_ '
								' WITH '
								'	item, prior_ancestors '
							)
						else:  # assign sample
							# these must be detached and reattached to field ItemSamples container
							statement += (
								' WITH item, prior_ancestors '
								' MATCH (item)-[from: FROM]->(:ItemSamples) '
								'	-[:FROM]->(:Block)-[:IS_IN]->(:FieldBlocks)-[:IS_IN]->(field:Field) '
								' DELETE from '
								' MERGE (is:ItemSamples)-[:FROM]->(field) '
								' CREATE (item)-[:FROM]->(is) '
								' WITH DISTINCT '
								'	item, prior_ancestors '
							)
					else:  # to tree or sample (these are all samples) and the values are lists
						# for these we need to find the most recent record and update accordingly,
						# if no record then we re-attach to the Field ItemSamples
						# we only need to assess one record since if they are submitted concurrently they must agree
						# we also don't need to check for variety conflicts as the update will be to
						# either the same items, a subset thereof or an ancestor
						statement += (
							' WITH '
							'	item, '
							'	prior_ancestors, '
							'	collect([input.name_lower, record.value, s.time]) as records, '
							'	max(s.time) as latest '
							' WITH item, prior_ancestors, [x IN records WHERE x[2] = latest][0] as record '
							' WITH item, prior_ancestors, record[0] as input_name_lower, record[1] as value'
							' MATCH '
							'	(item)-[: IS_IN | FROM *]->(field: Field) '
							' OPTIONAL MATCH '
							'	(new_source: Item)-[: IS_IN | FROM *]->(field) '
							'	WHERE '
							'		any(x IN labels(new_source) WHERE input_name_lower CONTAINS ("assign sample to " + toLower(x))) '
							'	AND '
							'		CASE '
							'			WHEN '
							'				input_name_lower CONTAINS "by name" THEN new_source.name_lower IN value '
							'			WHEN '
							'				input_name_lower CONTAINS "by id" THEN new_source.id IN value '
							'		END '
							' WITH item, prior_ancestors, collect(new_source) as new_sources, field '
							' WITH '
							'	item, '
							'	prior_ancestors, '
							'	CASE '
							'		WHEN length(new_sources) = 0 THEN [field] '
							'		ELSE new_sources '
							'		END '	
							'	as new_sources '
							' UNWIND new_sources AS new_source '
							' FOREACH (n IN CASE WHEN NOT "Sample" IN labels(new_source) THEN [1] ELSE [] END | '
							'	MERGE (:ItemSamples)-[:FROM]->(new_source) '
							' ) '
							' WITH item, prior_ancestors, new_source '
							' OPTIONAL MATCH (item_samples: ItemSamples)-[:FROM]->(new_source) '
							' WITH item, prior_ancestors, collect(coalesce(item_samples, new_source)) as new_sources '
							' MATCH (item)-[from: FROM]->() '
							' DELETE from '
							' WITH DISTINCT item, prior_ancestors, new_sources '
							' UNWIND new_sources as new_source '
							' MERGE '
							'	(item)-[:FROM]->(new_source) '
							' WITH distinct item, prior_ancestors '
						)
					statement += (
						' OPTIONAL MATCH '
						'	(item)-[:FROM | IS_IN*]->(ancestor: Item) '
						' WITH '
						'	item, '
						'	prior_ancestors, '
						'	collect(ancestor) as ancestors '
						' WITH '
						'	item, '
						'	ancestors,'
						'	[x IN prior_ancestors WHERE NOT x IN ancestors | x ] as removed_ancestors'
						' OPTIONAL MATCH '
						'	(item)<-[:FROM | IS_IN *]-(descendant: Item) '
						' WITH '
						'	item, '
						'	removed_ancestors, '
						'	ancestors + collect(descendant) as lineage '
						' UNWIND lineage AS kin '
						'	OPTIONAL MATCH '
						'		(kin)-[:FROM | IS_IN*]->(kin_ancestor: Item) '
						'	WITH '
						'		item, removed_ancestors,'
						'		kin,'
						'		collect(kin_ancestor) as kin_ancestors '
						'	OPTIONAL MATCH '
						'		(kin)<-[:FROM | IS_IN*]-(kin_descendant: Item) '
						'	WITH '
						'		item, removed_ancestors,'
						'		kin, '
						'		kin_ancestors + collect(kin_descendant) as kin_lineage '
						'	UNWIND kin_lineage AS kin_of_kin '
						'	WITH '
						'		item, removed_ancestors, '
						'		kin, collect(distinct kin_of_kin.variety) as kin_varieties '
						'	SET kin.varieties = CASE WHEN kin.variety IS NOT NULL THEN [kin.variety] ELSE kin_varieties END '
						' WITH '
						'	item, removed_ancestors, '
						'	collect(distinct kin.variety) as varieties '
						' SET item.varieties = CASE WHEN item.variety IS NOT NULL THEN [item.variety] ELSE varieties END '
						' WITH item, removed_ancestors '
						' UNWIND removed_ancestors AS removed_ancestor '
						'	OPTIONAL MATCH '
						'		(removed_ancestor)-[: FROM | IS_IN]->(descendant: Item) '
						'	WITH '
						'		item, removed_ancestor, '
						'		collect(descendant) as removed_ancestor_descendants '
						'	OPTIONAL MATCH '
						'		(removed_ancestor)<-[:FROM | IS_IN]-(ancestor: Item) '
						'	WITH '
						'		item, removed_ancestor, '
						'		removed_ancestor_descendants + collect(ancestor) as removed_ancestor_lineage '
						'	UNWIND removed_ancestor_lineage AS removed_ancestor_kin '
						'	WITH '
						'		item, removed_ancestor, '
						'		collect(removed_ancestor_kin.variety) as removed_ancestor_varieties '
						'	SET removed_ancestor.varieties = '
						'		CASE '
						'			WHEN removed_ancestor.variety IS NOT NULL THEN [removed_ancestor.variety] '
						'			ELSE removed_ancestor_varieties '
						'		END '
					)
					tx.run(
						statement,
						uid_list=property_uid[input_variable],
						property_inputs=properties_inputs[inputs_properties[input_variable]]
					)
				elif inputs_properties[input_variable] == 'time':
					# There is only one input for date and possibly one for time for each level of item.
					# so we always update on delete and can just check the set values
					statement = (
						' UNWIND $uid_list as uid'
						' MATCH '
						'	(item: Item {uid: uid}) '
					)
					if 'date' in input_variable:
						statement += (
							' REMOVE item.date, item.time '
						)
					else:
						statement += (
							' REMOVE item.time_of_day'
							'	 SET item.time = CASE '
							'		WHEN item.date IS NOT NULL '
							'		THEN '
							'			apoc.date.parse(item.date + " 12:00", "ms", "yyyy-MM-dd HH:mm") '
							'		END '
						)
					tx.run(
						statement,
						uid_list=property_uid[input_variable],
					)
				elif inputs_properties[input_variable] == 'elevation':
					tx.run(
						' UNWIND $uid_list as uid'
						' MATCH '
						'	(item: Item {uid: uid}) '
						' REMOVE item.elevation ',
						uid_list=property_uid[input_variable]
					)

	@celery.task(bind=True)
	def async_correct(self, username, access, upload_object):
		try:
			upload_object.set_fieldnames()
			if upload_object.error_messages:
				with app.app_context():
					error_messages = '<br>'.join(upload_object.error_messages)
					html = render_template(
						'emails/upload_report.html',
						response=(
							'Submission report for file: ' + upload_object.raw_filename + '\n<br>' +
							error_messages
						)
					)
					subject = "BreedCAFS upload rejected"
					recipients = [User(username).find('')['email']]
					response = (
							'Submission report for file: ' + upload_object.raw_filename + '\n<br>' +
							"Submission rejected due to invalid file:\n " + error_messages
					)
					body = response
					send_email(subject, app.config['ADMINS'][0], recipients, body, html)
				return {
					'status': 'ERRORS',
					'result': (
							'Submission report for file: ' + upload_object.raw_filename + '\n<br>' +
							error_messages

					)
				}
			with get_driver().session() as neo4j_session:
				with neo4j_session.begin_transaction() as tx:
					# todo Stop trimming before parsing, this should be done in one pass of the file
					if upload_object.file_extension == 'xlsx':
						wb = load_workbook(upload_object.file_path, read_only=True, data_only=True)
						if 'Records' not in wb.sheetnames:
							return {
								'status': 'ERRORS',
								'result': (
										'Submission report for file: ' + upload_object.raw_filename + '\n<br>' +
										'This workbook does not contain the expected "Records" worksheet'
								)
							}
					for worksheet in list(upload_object.fieldnames.keys()):
						if upload_object.file_extension == 'xlsx':
							if worksheet != 'Records':
								continue
						upload_object.trim_file(worksheet)
						if not upload_object.row_count[worksheet]:
							upload_object.error_messages.append('No records found to delete')
						upload_object.parse_rows(worksheet)
						upload_object.db_check(tx, worksheet)
						if upload_object.error_messages:
							error_messages = '<br>'.join(upload_object.error_messages)
							with app.app_context():
								response = (
										'Submission report for file: ' + upload_object.raw_filename + '\n<br>' +
										error_messages
								)
								html = render_template(
									'emails/upload_report.html',
									response=response
								)
								subject = "BreedCAFS correction rejected"
								recipients = [User(username).find('')['email']]
								body = response
								send_email(subject, app.config['ADMINS'][0], recipients, body, html)
							return {
								'status': 'ERRORS',
								'result': response
							}
						deletion_result = upload_object.correct(tx, access, worksheet)
						# update properties where needed
						property_uid = {
							'set custom name': [],
							'set row': [],
							'set column': [],
							'set sample unit': [],
							'assign tree to block by name': [],
							'assign tree to block by id': [],
							'assign sample to block(s) by name': [],
							'assign sample to block(s) by id': [],
							'assign sample to tree(s) by id': [],
							'assign sample to sample(s) by id': [],
							'set harvest date': [],
							'set planting date': [],
							'set harvest time': [],
							'assign variety name': [],
							'assign variety (el frances code)': []
						}
						# just grab the UID where records are deleted and do the update from any remaining records
						record_tally = {}
						missing_row_indexes = []
						expected_row_index = 1
						if not deletion_result.peek():
							missing_row_indexes += list(range(2, upload_object.row_count[worksheet] + 2))
						else:
							for record in deletion_result:
								while (
										expected_row_index != record[0]['row_index']
										and expected_row_index <= upload_object.row_count[worksheet] + 2
								):
									expected_row_index += 1
									if expected_row_index != record[0]['row_index']:
										missing_row_indexes.append(expected_row_index)
								if not record[0]['input variable'] in record_tally:
									record_tally[record[0]['input variable']] = 0
								record_tally[record[0]['input variable']] += 1
								if record[0]['input variable'] in property_uid:
									property_uid[record[0]['input variable']].append(record[0]['uid'])
							upload_object.remove_properties(tx, property_uid)
							if not expected_row_index == upload_object.row_count[worksheet]:
								missing_row_indexes += list(range(expected_row_index, upload_object.row_count[worksheet] + 2))
						if missing_row_indexes:
							missing_row_ranges = (
								list(x) for _, x in itertools.groupby(enumerate(missing_row_indexes), lambda i_x: i_x[0]-i_x[1])
							)
							missing_row_str = str(
								",".join("-".join(map(str, (i[0][1], i[-1][1])[:len(i)])) for i in missing_row_ranges)
							)
							tx.rollback()
							with app.app_context():
								# send result of merger in an email
								subject = 'BreedCAFS correction rejected'
								recipients = [User(username).find('')['email']]
								response = (
									'Submission report for file: ' + upload_object.raw_filename + '\n<br>' +
									'Correction rejected:\n '
								)
								if missing_row_str:
									response += (
										'<p>Records from the following rows of the uploaded file were not found: '
										+ missing_row_str
										+ '\n</p>'
									)
								body = response
								html = render_template(
									'emails/upload_report.html',
									response=response
								)
								send_email(subject, app.config['ADMINS'][0], recipients, body, html)
							return {
								'status': 'ERRORS',
								'result': response
							}
						tx.commit()
						with app.app_context():
							# send result of merger in an email
							subject = 'BreedCAFS correction summary'
							recipients = [User(username).find('')['email']]
							response = (
								'Submission report for file: %s \n<br>'
								'Correction report:\n '
								% upload_object.raw_filename
							)
							if record_tally:
								response += '<p>The following records were deleted: \n</p>'
								for key in record_tally:
									response += '<p>  -' + str(record_tally[key]) + ' ' + str(key) + ' records deleted\n </p>'
							body = response
							html = render_template(
								'emails/upload_report.html',
								response=response
							)
							send_email(subject, app.config['ADMINS'][0], recipients, body, html)
						return {
							'status': 'SUCCESS',
							'result': response
							}
		except (ServiceUnavailable, SecurityError) as exc:
			raise self.retry(exc=exc)


class Upload:
	@classmethod
	def select(cls, submission_type, filename):
		type_to_class_map = {
			'table.csv': CSVTable
		}
		file_extension = os.path.splitext(filename)[1]
		upload_type = ''.join(submission_type, file_extension)
		if upload_type not in type_to_class_map:
			raise ValueError('Unrecognised file type for upload')

		return type_to_class_map[upload_type]

	def __init__(self, username, raw_filename):
		self.username = username
		self.raw_filename = raw_filename
		self.filename = secure_filename(raw_filename)
		self.file_path = os.path.join(app.config['UPLOAD_FOLDER'], username, self.filename)
		self.file_extension = os.path.splitext(self.filename)[1]
		self.error_messages = []


class CSVTable(Upload):
	def __init__(self, username, raw_filename):
		super().__init__(username, raw_filename)
		self.fieldnames = None
		self.fieldnames_lower = None
		self.input_variables = None
		self.record_type = None
		self.row_fields = None
		self.value_fields = None
		self.records = None
		self.file_errors = set()
		self.error_rows = {} # row indexed row dicts where row contain errors
		self.row_errors = {} # row and field indexed error messages

	def add_row_error(self, row_index, field, message):
		if row_index not in self.row_errors:
			self.row_errors[row_index] = {}
		if field not in self.row_errors[row_index]:
			self.row_errors[row_index][field] = set()
		self.row_errors[row_index][field].add(message)

	def check_format(self):
		with open(self.file_path) as csv_file:
			self.check_format(csv_file)
			self.parse_fieldnames(csv_file)

	def check_format(self, csv_file):
		# check basic csv formatting
		# TODO implement CSV kit checks
		# now get the dialect and check it conforms to expectations
		dialect = csv.Sniffer().sniff(csv_file.read())
		if not all((dialect.delimiter == ',', dialect.quotechar == '"')):
			logging.info(
				'Bad csv file submitted: \n'
				'username: %s'
				'filename: %s' % (
					self.username,
					self.file_path
				)
			)
			self.file_errors.add('.csv files must be comma (,) separated with double quoted (") fields')

	def parse_fieldnames(self, csv_file):
		self.set_fieldnames(csv_file)
		self.get_input_variables()
		self.sort_fieldnames()


	def set_fieldnames(self, csv_file):
		if self.file_errors:
			pass
		reader = csv.reader(csv_file)
		# check fieldnames for duplicates
		self.fieldnames = next(reader)
		self.fieldnames_lower = [fieldname.lower().strip() for fieldname in self.fieldnames if fieldname.strip()]
		if len(self.fieldnames_lower) > len(set(self.fieldnames)):
			logging.info(
				'Duplicate or whitespace column labels in submitted csv table: \n'
				'username: %s '
				'filename: %s' % (
					self.username,
					self.file_path
				)
			)
			self.file_errors.add(
				'CSV tables with duplicate (or whitespace) column labels are not supported (case is ignored).'
			)

	def get_input_variables(self):
		if self.file_errors:
			pass
		query = (
			' MATCH '
			'	(input: Input)-[:OF_TYPE]->(record_type:RecordType) '
			' WHERE '
			'	record_type.name_lower IN ["property", "trait", "condition"] '  
			'	AND input.name_lower IN $fieldnames_lower'
			' RETURN {'
			'	record_type: record_type.name_lower, '
			'	name: input.name_lower, '
			'	format: input.format, '
			'	categories = input.category_list '
			' } '
		)
		parameters = {
			'fieldnames_lower': self.fieldnames_lower
		}
		input_variables = Query().list_result(query, parameters)
		if not input_variables:
			query = (
				' MATCH '
				'	(input: Input)-[:OF_TYPE]->(:RecordType {name_lower:"curve"}) '
				' WHERE toLower($filename) CONTAINS input.name_lower'
				' RETURN {'
				'	record_type: record_type.name_lower, '
				'	name: input.name_lower, '
				'	format: input.format, '
				'	categories = input.category_list '
				' } '
			)
			parameters = {
				'filename': self.raw_filename
			}
			input_variables = [Query().bolt_result(query, parameters).single()[0]]
		if not input_variables:
			logging.info(
				'CSV Table file submitted without recognised input variables: \n'
				'username: %s '
				'filename: %s ' % (
					self.username,
					self.file_path
				)
			)
			self.file_errors.add(
				'This file does not contain recognised input variables. '
				'BreedCAFS registered input variables must be in column headers '
				'(for property, trait and condition record types) '
				'or contained in the filename (for curve record types)'
			)
			pass
		record_types = [input_variable['record_type'] for input_variable in input_variables]
		if len(record_types) > 1:
			logging.info(
				'CSV Table file submitted with multiple record types: \n'
				'username: %s '
				'filename: %s ' % (
					self.username,
					self.file_path
				)
			)
			self.file_errors.add(
				'This .csv file contains input variables of more than one record type. This is not supported.'
			)
			pass
		self.record_type = record_types[1]
		self.input_variables = {}
		for v in input_variables:
			self.input_variables[v['name']] = {
				'format': v['format'],
				'categories': v['categories']
			}

	def sort_fieldnames(self):
		if self.file_errors:
			pass
		if not app.config['REQUIRED_FIELDNAMES'][self.record_type].issubset(self.fieldnames_lower):
			self.file_errors.add(
				'The required fieldnames of the %s record type are: %s' %
				(
					self.record_type,
					', '.join(app.config['RECORD_TYPE_REQUIRED'][self.record_type]),
				)
			)
			pass
		row_fields = app.config['REQUIRED_FIELDNAMES'][self.record_type] | app.config['OPTIONAL_FIELDNAMES'[self.record_type]]
		fieldnames = set(self.fieldnames_lower)
		self.row_fields = fieldnames.intersection(row_fields)
		value_fields = fieldnames - self.row_fields - app.config['REFERENCE_FIELDNAMES']
		if self.record_type == 'curve':
			all_numbers = True
			for fieldname in value_fields:
				try:
					float(fieldname)
				except ValueError:
					all_numbers = False
					break
			if all_numbers:
				self.value_fields = value_fields
		else:
			input_variables = set(self.input_variables.keys())
			if not input_variables == value_fields:
				self.file_errors.add(
					'Unrecognised fieldnames for %s records: %s' % (
						self.record_type,
						', '.join(value_fields - input_variables)
					)
				)
				pass

	def batch_process(self, n_rows = 100):
		with open(self.file_path) as csv_file:
			# enumerate the generator skipping header and using 1 based row index as this is more familiar to users
			# also reading in with lowercase fieldnames for more efficient matching of keys
			dict_reader = enumerate(csv.DictReader(csv_file, fieldnames = self.fieldnames_lower), start = 2)
			self.records = {input_variable: [] for input_variable in self.input_variables.keys()}
			record_class = Record.select(self.record_type)
			while True:
				try:
					self.parse_batch(dict_reader, n_rows, record_class)
					if self.row_errors:
						return # ERROR TABLE
					self.submit_batch()
				except StopIteration:
					break

	def parse_batch(self, dict_reader, n_rows, record_class):
		try:
			errors = False
			row_count = 0
			keep_going = True
			row_parser = TableRowParser(
				self.record_type,
				self.input_variables,
				self.row_fields,
				self.value_fields,
				self.row_errors
			)
			while keep_going:
				# parse rows until the row limit unless there are errors
				# in which case keep going until the error row limit is reached
				row_count += 1
				if row_count > n_rows:
					if not errors or len(self.error_rows.keys()) > app.config['BATCH_PROCESS_MAX_ERROR_ROWS']:
						keep_going = False
				try:
					row_index, row_dict = next(dict_reader)
				except StopIteration:
					raise
				# First check the row reference detail formatting, e.g. uid
				row_parser.check_row_fields(row_index, row_dict)
				if self.row_errors[row_index]:
					continue
				# then create records
				for input_variable in self.input_variables.keys():
					if self.record_type == 'curve':
						record = record_class(row_index, row_dict, self.row_errors).create(
							self.value_fields,
							row_dict
						)
					else:
						record = record_class(row_index, row_dict, self.row_errors).create(
							input_variable,
							row_dict[input_variable],
							self.input_variables[input_variable]['format'],
							self.input_variables[input_variable]['categories']
						)
					if not self.row_errors:  # only keep records if no errors in batch
						self.records[input_variable].append(record)
		except StopIteration:
			raise

	def error_table(self):
		# create a html table string with tooltip for details
		if self.error_rows:
			error_table_head = '<tr><th>Line</th><th>' + '</th><th>'.join(self.fieldnames) + '</th></tr>\n'
			for row_index, row_dict in self.error_rows:
				table_row = '<tr><td>' + str(row_index) + '</td><td '
				errors = None
				if self.record_type == 'curve':
					errors = self.error_records[row_index][self.input_variables[0]]
				for fieldname_lower in self.fieldnames_lower:
					if not errors:
						errors = self.error_records[row_index[fieldname_lower]]



					:
						table_row += 'bgcolor = #FFFF00 title = "' + '/n - '.join(record.errors[fieldname_lower])
					table_row += '><p> ' + self.error_rows[record.row_index][fieldname_lower] + '</p></td>'
				table_row += '<tr>'



		table = <table>
			for key, value in record.errors:
			   <tr>
					<th> {{ key }} </th>
					<td> {{ value }} </td>
			   </tr>
			{% endfor %}
			</table>

		if self.errors:
			header_string = '<tr><th><p>Line#</p></th>'
			for field in self.fieldnames:
				if self.field_errors:
					if field in self.field_errors:
						header_string += '<th bgcolor = #FFFF00 title = "' + self.field_errors[field] \
										 + '"><p>' + field + '</p></th>'
					else:
						header_string += '<th><p>' + field + '</p></th>'
				else:
					header_string += '<th><p>' + field + '</p></th>'
			html_table = header_string + '</tr>'
			# construct each row and append to growing table
			for i, item in enumerate(sorted(self.errors)):
				if i >= max_length:
					return '<div id="response_table_div"><table>' + html_table + '</table></div>'
				row_string = self.errors[item].html_row(self.fieldnames)
				html_table += row_string
			return '<div id="response_table_div"><table>' + html_table + '</table></div>'
		else:
			return None


class UploadCSVDB:
	if self.submission_type in ['internal_db', 'db']:
		self.record_type = 'mixed'
		if not app.config['RECORD_TYPE_REQUIRED'][self.record_type].issubset(self.fieldnames):
			logging.info(
				'Missing required fieldnames in submitted file: \n'
				+ 'username: ' + self.username
				+ 'filename: ' + self.file_path
			)
			self.error_messages.append(
				'This file does not appear to be a database exported csv file. ' +
				'It should contain the following fieldnames: ' +
				', '.join([str(i) for i in app.config['RECORD_TYPE_REQUIRED'][self.record_type]])
			)
















class UploadXLSX:
	def __init__(self, username, file_path):
		self.username = username
		self.file_path = file_path
		self.error_messages = []
		self.workbook = None

	def format_check(self):
		# check the worksheet names and return exception if not a valid xlsx file
		try:
			self.workbook = load_workbook(self.file_path, read_only=True, data_only=True)
			if not set(self.workbook.sheetnames).issubset(set(app.config['WORKSHEET_NAMES'].values())):
				unidentified_worksheets = set(self.workbook.sheetnames) - set(app.config['WORKSHEET_NAMES'].values())
				# Check other worksheet names to see if they match with curve input names
				statement = (
					' MATCH '
					'	(input: Input)-[:OF_TYPE]->(:RecordType {name_lower:"curve"}) '
					' WHERE input.name_lower IN $names '
					' RETURN count(input) '
				)
				parameters = {
					'names': [i.lower for i in unidentified_worksheets]
				}
				result = Query().bolt_result(statement, parameters).single()
				if not result[0] == len(unidentified_worksheets):
					logging.info(
						'.xlsx file submitted without valid worksheet names: \n'
						+ 'username: ' + self.username
						+ 'filename: ' + self.file_path
					)
					self.error_messages.append(
						'This workbook contains unrecognised worksheets'
					)
		except BadZipfile:
			logging.info(
				'Bad zip file submitted: \n'
				+ 'username: ' + self.username
				+ 'filename: ' + self.file_path
			)
			self.error_messages.append('The uploaded file does not appear to be a valid xlsx workbook.')